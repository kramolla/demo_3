from db.connection import Base, engine, SessionLocal
from openpyxl import load_workbook
from models.product_type import ProductTypeModel
from models.product import ProductModel
from models.partner import PartnerModel
from models.order import OrderModel
from models.material import MaterialModel

# Пути к файлам Excel
BASE_DIR = "excel/"
product_type_file = BASE_DIR + "Product_type_import.xlsx"
products_file = BASE_DIR + "Products_import.xlsx"
partners_file = BASE_DIR + "Partners_import.xlsx"
sales_history_file = BASE_DIR + "Partner_products_import.xlsx"
materials_file = BASE_DIR + "Material_type_import.xlsx"

#Удаляет существующие таблицы, использовать не обязательно
Base.metadata.drop_all(engine)
#Создаёт таблицы
Base.metadata.create_all(engine)
session = SessionLocal()

try:
    # Импорт типов продукции
    wb = load_workbook(product_type_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        production_type, coefficient = row

        # Пропускаем строки с пустыми значениями
        if not production_type or not coefficient:
            continue

        session.add(ProductTypeModel(
            production_type=production_type,
            coefficient=coefficient
        ))
    session.commit()
    print("Типы продукции добавлены.")

    # Импорт продуктов
    wb = load_workbook(products_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_type_name, name, article_number, minimum_cost = row
        product_type = session.query(ProductTypeModel).filter_by(production_type=product_type_name).first()
        session.add(ProductModel(
            product_type_id=product_type.id,
            name=name,
            article_number=article_number,
            minimum_cost=minimum_cost))
    session.commit()
    print("Продукты добавлены.")

    # Импорт партнёров
    wb = load_workbook(partners_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        partner_type, name, director, email, phone, legal_address, inn, rating = row
        session.add(PartnerModel(
            partner_type=partner_type,
            name=name,
            director=director,
            email=email,
            phone=phone,
            legal_address=legal_address,
            inn=inn,
            rating=rating
        ))
    session.commit()
    print("Партнёры добавлены.")

    # Импорт истории продаж
    wb = load_workbook(sales_history_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, partner_name, quantity, sale_date = row
        product = session.query(ProductModel).filter_by(name=product_name).first()
        partner = session.query(PartnerModel).filter_by(name=partner_name).first()
        session.add(OrderModel(
            partner_id=partner.id,
            product_id=product.id,
            quantity=quantity,
            sale_date=sale_date
        ))
    session.commit()
    print("История продаж добавлена.")

    # Импорт материалов
    wb = load_workbook(materials_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        material_type, percentage_of_defective_material = row
        session.add(MaterialModel(
            type=material_type,
            percentage_of_defective_material=percentage_of_defective_material
        ))
    session.commit()
    print("Материалы добавлены.")

# Реакция программы на ошибку
except Exception as e:
    session.rollback()
    print(f"Ошибка при импорте данных: {e}")
# Код, который запуститься вне зависимости от успешности выполнения программы
finally:
    session.close()
    print("Импорт данных завершён.")
